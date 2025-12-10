import smtplib
import ssl
import json
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from email import encoders
from typing import Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.config import settings


class EmailService:
    def __init__(self):
        self.smtp_server = settings.smtp_server
        self.smtp_port = settings.smtp_port
        self.username = settings.smtp_username
        self.password = settings.smtp_password
        self.support_email = settings.company_support_email

    def send_invoice_data(
        self,
        recipient_email: str,
        invoice_data: Dict[str, Any],
        image_filename: str,
        user_email: str
    ) -> bool:
        """Send invoice data via email"""
        try:
            # Create message container
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"Invoice Data - {image_filename}"
            msg['From'] = self.support_email
            msg['To'] = recipient_email

            # Create HTML content
            html_content = self._create_invoice_html(invoice_data, image_filename, user_email)
            
            # Create plain text version
            text_content = self._create_invoice_text(invoice_data, image_filename, user_email)

            # Add both versions
            text_part = MIMEText(text_content, 'plain')
            html_part = MIMEText(html_content, 'html')
            
            msg.attach(text_part)
            msg.attach(html_part)

            # Send email
            return self._send_email(msg, recipient_email)

        except Exception as e:
            print(f"Error sending invoice email: {str(e)}")
            return False

    def send_invoice_json(
        self,
        recipient_email: str,
        invoice_data: Dict[str, Any],
        image_filename: str,
        user_email: str
    ) -> bool:
        """Send invoice data as JSON attachment"""
        try:
            # Create message container
            msg = MIMEMultipart()
            msg['Subject'] = f"Invoice Data (JSON) - {image_filename}"
            msg['From'] = self.support_email
            msg['To'] = recipient_email

            # Email body
            body = f"""
Hello,

Please find attached the extracted invoice data for {image_filename}.

This data was processed and sent by user: {user_email}
Processing date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}

Best regards,
DoxSnap Team
{self.support_email}
"""
            msg.attach(MIMEText(body, 'plain'))

            # Create JSON attachment
            json_data = json.dumps(invoice_data, indent=2, ensure_ascii=False)
            attachment = MIMEBase('application', 'json')
            attachment.set_payload(json_data.encode('utf-8'))
            encoders.encode_base64(attachment)
            
            filename = f"{image_filename}_data.json"
            attachment.add_header(
                'Content-Disposition',
                f'attachment; filename="{filename}"'
            )
            msg.attach(attachment)

            # Send email
            return self._send_email(msg, recipient_email)

        except Exception as e:
            print(f"Error sending invoice JSON: {str(e)}")
            return False

    def send_invoice_excel(
        self,
        recipient_email: str,
        invoice_data_list: List[Dict[str, Any]],
        user_email: str
    ) -> bool:
        """Send all processed documents as Excel attachment with Header and Details sheets"""
        try:
            # Create message container
            msg = MIMEMultipart()
            msg['Subject'] = f"All Processed Documents - Excel Report"
            msg['From'] = self.support_email
            msg['To'] = recipient_email

            # Email body
            body = f"""
Hello,

Please find attached an Excel report containing all your processed documents.

The Excel file contains two sheets:
- Header: Summary information for each document  
- Details: Complete line items and details for all invoices

📊 Report Details:
- Total documents included: {len(invoice_data_list)}
- Processed by user: {user_email}
- Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}
- File format: Microsoft Excel (.xlsx)

📎 Attachment Information:
- File name: processed_documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx
- If you don't see the attachment, please check your spam/junk folder
- The attachment should appear as an Excel file that you can download and open

Best regards,
DoxSnap Team
{self.support_email}
"""
            msg.attach(MIMEText(body, 'plain'))

            # Create Excel file in memory
            excel_data = self._create_excel_file(invoice_data_list)
            
            # Create Excel attachment using standard octet-stream for better compatibility
            filename = f"processed_documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            attachment = MIMEApplication(excel_data, name=filename)
            attachment.add_header(
                'Content-Disposition',
                f'attachment; filename="{filename}"'
            )
            # Set explicit content type for Excel files
            attachment.replace_header('Content-Type', 'application/vnd.ms-excel')
            msg.attach(attachment)

            # Send email
            return self._send_email(msg, recipient_email)

        except Exception as e:
            print(f"Error sending invoice Excel: {str(e)}")
            return False

    def _create_excel_file(self, invoice_data_list: List[Dict[str, Any]]) -> bytes:
        """Create Excel file with Header and Details sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Header sheet
        header_sheet = wb.create_sheet("Header")
        self._create_header_sheet(header_sheet, invoice_data_list)
        
        # Create Details sheet
        details_sheet = wb.create_sheet("Details")
        self._create_details_sheet(details_sheet, invoice_data_list)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()

    def _create_header_sheet(self, sheet, invoice_data_list: List[Dict[str, Any]]):
        """Create header sheet with summary information"""
        
        # Set up header row
        headers = [
            'Document ID', 'Filename', 'Processing Date', 'Invoice Number', 
            'Invoice Date', 'Due Date', 'Supplier Company', 'Customer Company',
            'Subtotal', 'Tax Amount', 'Total Amount', 'Processing Method', 
            'Confidence Score', 'Words Extracted', 'OCR Confidence'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row, invoice_data in enumerate(invoice_data_list, 2):
            structured_data = invoice_data.get('structured_data', {})
            document_info = structured_data.get('document_info', {})
            supplier = structured_data.get('supplier', {})
            customer = structured_data.get('customer', {})
            financial_details = structured_data.get('financial_details', {})
            ocr_stats = invoice_data.get('ocr_stats', {})
            
            # Fill data
            sheet.cell(row=row, column=1, value=invoice_data.get('image_id', ''))
            sheet.cell(row=row, column=2, value=invoice_data.get('original_filename', ''))
            sheet.cell(row=row, column=3, value=invoice_data.get('created_at', ''))
            sheet.cell(row=row, column=4, value=document_info.get('invoice_number', ''))
            sheet.cell(row=row, column=5, value=document_info.get('invoice_date', ''))
            sheet.cell(row=row, column=6, value=document_info.get('due_date', ''))
            sheet.cell(row=row, column=7, value=supplier.get('company_name', ''))
            sheet.cell(row=row, column=8, value=customer.get('company_name', ''))
            sheet.cell(row=row, column=9, value=financial_details.get('subtotal', 0))
            sheet.cell(row=row, column=10, value=financial_details.get('total_tax_amount', 0))
            sheet.cell(row=row, column=11, value=financial_details.get('total_after_tax', 0))
            sheet.cell(row=row, column=12, value=invoice_data.get('processing_method', ''))
            sheet.cell(row=row, column=13, value=invoice_data.get('extraction_confidence', 0))
            sheet.cell(row=row, column=14, value=ocr_stats.get('words_extracted', 0))
            sheet.cell(row=row, column=15, value=ocr_stats.get('average_confidence', 0))
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_details_sheet(self, sheet, invoice_data_list: List[Dict[str, Any]]):
        """Create details sheet with line items"""
        
        # Set up header row
        headers = [
            'Document ID', 'Filename', 'Invoice Number', 'Line Item #', 
            'Item Description', 'Quantity', 'Unit Price', 'Line Total',
            'Supplier Company', 'Customer Company', 'Invoice Date'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        current_row = 2
        for invoice_data in invoice_data_list:
            structured_data = invoice_data.get('structured_data', {})
            document_info = structured_data.get('document_info', {})
            supplier = structured_data.get('supplier', {})
            customer = structured_data.get('customer', {})
            line_items = structured_data.get('line_items', [])
            
            if line_items:
                for line_num, item in enumerate(line_items, 1):
                    sheet.cell(row=current_row, column=1, value=invoice_data.get('image_id', ''))
                    sheet.cell(row=current_row, column=2, value=invoice_data.get('original_filename', ''))
                    sheet.cell(row=current_row, column=3, value=document_info.get('invoice_number', ''))
                    sheet.cell(row=current_row, column=4, value=line_num)
                    sheet.cell(row=current_row, column=5, value=item.get('description', ''))
                    sheet.cell(row=current_row, column=6, value=item.get('quantity', 0))
                    sheet.cell(row=current_row, column=7, value=item.get('unit_price', 0))
                    sheet.cell(row=current_row, column=8, value=item.get('total_line_amount', 0))
                    sheet.cell(row=current_row, column=9, value=supplier.get('company_name', ''))
                    sheet.cell(row=current_row, column=10, value=customer.get('company_name', ''))
                    sheet.cell(row=current_row, column=11, value=document_info.get('invoice_date', ''))
                    current_row += 1
            else:
                # Add a row even if no line items
                sheet.cell(row=current_row, column=1, value=invoice_data.get('image_id', ''))
                sheet.cell(row=current_row, column=2, value=invoice_data.get('original_filename', ''))
                sheet.cell(row=current_row, column=3, value=document_info.get('invoice_number', ''))
                sheet.cell(row=current_row, column=4, value='N/A')
                sheet.cell(row=current_row, column=5, value='No line items found')
                sheet.cell(row=current_row, column=9, value=supplier.get('company_name', ''))
                sheet.cell(row=current_row, column=10, value=customer.get('company_name', ''))
                sheet.cell(row=current_row, column=11, value=document_info.get('invoice_date', ''))
                current_row += 1
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def send_email(self, recipient_email: str, subject: str, message: str) -> bool:
        """Send a simple email with HTML content"""
        try:
            # Create message container
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = self.support_email
            msg['To'] = recipient_email
            
            # Create HTML content with proper styling
            html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ background: linear-gradient(135deg, #4f46e5, #7c3aed); color: white; padding: 30px; text-align: center; border-radius: 8px 8px 0 0; }}
        .header h1 {{ margin: 0; font-size: 24px; }}
        .content {{ padding: 30px; }}
        .footer {{ background: #f8f9fa; padding: 20px; text-align: center; border-radius: 0 0 8px 8px; font-size: 12px; color: #666; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>DoxSnap</h1>
        </div>
        <div class="content">
            {message}
        </div>
        <div class="footer">
            <p>This email was sent from DoxSnap. For support, contact: {self.support_email}</p>
        </div>
    </div>
</body>
</html>
            """
            
            # Add HTML content
            html_part = MIMEText(html_content, 'html')
            msg.attach(html_part)
            
            # Send email using the internal method
            return self._send_email(msg, recipient_email)
            
        except Exception as e:
            print(f"Error sending email: {str(e)}")
            return False

    def _send_email(self, msg: MIMEMultipart, recipient_email: str) -> bool:
        """Send the email message using the same logic as send_email.py"""
        try:
            # Validate configuration
            if not self.username or not self.password:
                print("Email configuration incomplete: missing username or password")
                return False
            
            # Use the same connection logic as the working send_email.py
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()  # Simple starttls without context like the working version
            server.login(self.username, self.password)
            
            # Send email
            text = msg.as_string()
            server.sendmail(self.support_email, recipient_email, text)
            server.quit()
            
            print(f"Email sent successfully to {recipient_email}")
            return True
            
        except smtplib.SMTPAuthenticationError as e:
            print(f"SMTP Authentication failed: {str(e)}")
            print("Possible solutions:")
            print("1. Enable 2-factor authentication and use an App Password")
            print("2. Check if 'Less secure app access' is enabled (not recommended)")
            print("3. Verify the email credentials are correct")
            return False
            
        except smtplib.SMTPException as e:
            print(f"SMTP error occurred: {str(e)}")
            return False
            
        except Exception as e:
            print(f"Failed to send email: {str(e)}")
            return False

    def _create_invoice_html(self, invoice_data: Dict[str, Any], filename: str, user_email: str) -> str:
        """Create HTML formatted invoice email"""
        
        # Extract structured data
        structured_data = invoice_data.get('structured_data', {})
        document_info = structured_data.get('document_info', {})
        supplier = structured_data.get('supplier', {})
        customer = structured_data.get('customer', {})
        financial_details = structured_data.get('financial_details', {})
        line_items = structured_data.get('line_items', [])
        
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Invoice Data - {filename}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #333; }}
        .header {{ background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
        .section {{ margin-bottom: 25px; }}
        .section-title {{ background-color: #007bff; color: white; padding: 10px; margin: 0; font-size: 16px; }}
        .section-content {{ border: 1px solid #dee2e6; padding: 15px; }}
        .info-row {{ margin-bottom: 8px; }}
        .label {{ font-weight: bold; color: #495057; }}
        .value {{ margin-left: 10px; }}
        .line-items {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        .line-items th {{ background-color: #f8f9fa; padding: 8px; border: 1px solid #dee2e6; text-align: left; }}
        .line-items td {{ padding: 8px; border: 1px solid #dee2e6; }}
        .footer {{ background-color: #f8f9fa; padding: 15px; border-radius: 8px; margin-top: 20px; font-size: 12px; color: #666; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Invoice Data Extraction</h1>
        <p><strong>File:</strong> {filename}</p>
        <p><strong>Processed by:</strong> {user_email}</p>
        <p><strong>Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
    </div>
"""

        # Document Information
        if document_info:
            html += """
    <div class="section">
        <h2 class="section-title">Document Information</h2>
        <div class="section-content">
"""
            if document_info.get('invoice_number'):
                html += f'<div class="info-row"><span class="label">Invoice Number:</span><span class="value">{document_info["invoice_number"]}</span></div>'
            if document_info.get('invoice_date'):
                html += f'<div class="info-row"><span class="label">Invoice Date:</span><span class="value">{document_info["invoice_date"]}</span></div>'
            if document_info.get('due_date'):
                html += f'<div class="info-row"><span class="label">Due Date:</span><span class="value">{document_info["due_date"]}</span></div>'
            html += "</div></div>"

        # Supplier Information
        if supplier:
            html += """
    <div class="section">
        <h2 class="section-title">Supplier Information</h2>
        <div class="section-content">
"""
            if supplier.get('company_name'):
                html += f'<div class="info-row"><span class="label">Company:</span><span class="value">{supplier["company_name"]}</span></div>'
            if supplier.get('company_address'):
                html += f'<div class="info-row"><span class="label">Address:</span><span class="value">{supplier["company_address"]}</span></div>'
            if supplier.get('phone'):
                html += f'<div class="info-row"><span class="label">Phone:</span><span class="value">{supplier["phone"]}</span></div>'
            if supplier.get('email'):
                html += f'<div class="info-row"><span class="label">Email:</span><span class="value">{supplier["email"]}</span></div>'
            html += "</div></div>"

        # Customer Information
        if customer:
            html += """
    <div class="section">
        <h2 class="section-title">Customer Information</h2>
        <div class="section-content">
"""
            if customer.get('contact_person'):
                html += f'<div class="info-row"><span class="label">Contact Person:</span><span class="value">{customer["contact_person"]}</span></div>'
            if customer.get('company_name'):
                html += f'<div class="info-row"><span class="label">Company:</span><span class="value">{customer["company_name"]}</span></div>'
            if customer.get('address'):
                html += f'<div class="info-row"><span class="label">Address:</span><span class="value">{customer["address"]}</span></div>'
            html += "</div></div>"

        # Line Items
        if line_items:
            html += """
    <div class="section">
        <h2 class="section-title">Line Items</h2>
        <div class="section-content">
            <table class="line-items">
                <thead>
                    <tr>
                        <th>Item</th>
                        <th>Description</th>
                        <th>Quantity</th>
                        <th>Unit Price</th>
                        <th>Total</th>
                    </tr>
                </thead>
                <tbody>
"""
            for i, item in enumerate(line_items, 1):
                description = item.get('description', f'Item {i}')
                quantity = item.get('quantity', 'N/A')
                unit_price = f"${item.get('unit_price', 0):.2f}" if item.get('unit_price') else 'N/A'
                total = f"${item.get('total_line_amount', 0):.2f}" if item.get('total_line_amount') else 'N/A'
                
                html += f"""
                    <tr>
                        <td>{i}</td>
                        <td>{description}</td>
                        <td>{quantity}</td>
                        <td>{unit_price}</td>
                        <td>{total}</td>
                    </tr>
"""
            html += "</tbody></table></div></div>"

        # Financial Summary
        if financial_details:
            html += """
    <div class="section">
        <h2 class="section-title">Financial Summary</h2>
        <div class="section-content">
"""
            if financial_details.get('subtotal'):
                html += f'<div class="info-row"><span class="label">Subtotal:</span><span class="value">${financial_details["subtotal"]:.2f}</span></div>'
            if financial_details.get('total_tax_amount'):
                html += f'<div class="info-row"><span class="label">Tax Amount:</span><span class="value">${financial_details["total_tax_amount"]:.2f}</span></div>'
            if financial_details.get('total_after_tax'):
                html += f'<div class="info-row"><span class="label"><strong>Total Amount:</strong></span><span class="value"><strong>${financial_details["total_after_tax"]:.2f}</strong></span></div>'
            html += "</div></div>"

        html += f"""
    <div class="footer">
        <p>This email was generated automatically by DoxSnap invoice processing system.</p>
        <p>For support, please contact: {self.support_email}</p>
    </div>
</body>
</html>
"""
        return html

    def _create_invoice_text(self, invoice_data: Dict[str, Any], filename: str, user_email: str) -> str:
        """Create plain text formatted invoice email"""
        
        structured_data = invoice_data.get('structured_data', {})
        document_info = structured_data.get('document_info', {})
        supplier = structured_data.get('supplier', {})
        customer = structured_data.get('customer', {})
        financial_details = structured_data.get('financial_details', {})
        line_items = structured_data.get('line_items', [])
        
        text = f"""
INVOICE DATA EXTRACTION
=======================

File: {filename}
Processed by: {user_email}
Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}

"""

        if document_info:
            text += "DOCUMENT INFORMATION\n" + "-" * 20 + "\n"
            if document_info.get('invoice_number'):
                text += f"Invoice Number: {document_info['invoice_number']}\n"
            if document_info.get('invoice_date'):
                text += f"Invoice Date: {document_info['invoice_date']}\n"
            if document_info.get('due_date'):
                text += f"Due Date: {document_info['due_date']}\n"
            text += "\n"

        if supplier:
            text += "SUPPLIER INFORMATION\n" + "-" * 20 + "\n"
            if supplier.get('company_name'):
                text += f"Company: {supplier['company_name']}\n"
            if supplier.get('company_address'):
                text += f"Address: {supplier['company_address']}\n"
            if supplier.get('phone'):
                text += f"Phone: {supplier['phone']}\n"
            if supplier.get('email'):
                text += f"Email: {supplier['email']}\n"
            text += "\n"

        if customer:
            text += "CUSTOMER INFORMATION\n" + "-" * 20 + "\n"
            if customer.get('contact_person'):
                text += f"Contact Person: {customer['contact_person']}\n"
            if customer.get('company_name'):
                text += f"Company: {customer['company_name']}\n"
            if customer.get('address'):
                text += f"Address: {customer['address']}\n"
            text += "\n"

        if line_items:
            text += "LINE ITEMS\n" + "-" * 10 + "\n"
            for i, item in enumerate(line_items, 1):
                text += f"Item {i}: {item.get('description', 'N/A')}\n"
                text += f"  Quantity: {item.get('quantity', 'N/A')}\n"
                text += f"  Unit Price: ${item.get('unit_price', 0):.2f}\n"
                text += f"  Total: ${item.get('total_line_amount', 0):.2f}\n\n"

        if financial_details:
            text += "FINANCIAL SUMMARY\n" + "-" * 17 + "\n"
            if financial_details.get('subtotal'):
                text += f"Subtotal: ${financial_details['subtotal']:.2f}\n"
            if financial_details.get('total_tax_amount'):
                text += f"Tax Amount: ${financial_details['total_tax_amount']:.2f}\n"
            if financial_details.get('total_after_tax'):
                text += f"TOTAL AMOUNT: ${financial_details['total_after_tax']:.2f}\n"
            text += "\n"

        text += f"""
---
This email was generated automatically by DoxSnap invoice processing system.
For support, please contact: {self.support_email}
"""

        return text