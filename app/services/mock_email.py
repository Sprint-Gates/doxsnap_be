"""
Mock email service for testing when SMTP is not available
This saves emails to files instead of sending them
"""

import json
import os
import io
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class MockEmailService:
    def __init__(self):
        # Create mock_emails directory
        self.output_dir = Path("mock_emails")
        self.output_dir.mkdir(exist_ok=True)

    def send_invoice_data(
        self,
        recipient_email: str,
        invoice_data: Dict[str, Any],
        image_filename: str,
        user_email: str
    ) -> bool:
        """Save invoice data as HTML file instead of sending email"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"invoice_html_{timestamp}_{recipient_email.replace('@', '_')}.html"
            filepath = self.output_dir / filename
            
            html_content = self._create_invoice_html(invoice_data, image_filename, user_email, recipient_email)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ Mock email saved to: {filepath}")
            print(f"   Recipient: {recipient_email}")
            print(f"   Invoice: {image_filename}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error saving mock email: {str(e)}")
            return False

    def send_invoice_json(
        self,
        recipient_email: str,
        invoice_data: Dict[str, Any],
        image_filename: str,
        user_email: str
    ) -> bool:
        """Save invoice data as JSON file instead of sending email"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"invoice_json_{timestamp}_{recipient_email.replace('@', '_')}.json"
            filepath = self.output_dir / filename
            
            # Create email metadata + invoice data
            email_data = {
                "email_info": {
                    "to": recipient_email,
                    "from": "noreply@coresrp.com",
                    "subject": f"Invoice Data (JSON) - {image_filename}",
                    "sent_by": user_email,
                    "timestamp": datetime.now().isoformat(),
                    "type": "json_attachment"
                },
                "invoice_data": invoice_data
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(email_data, f, indent=2, ensure_ascii=False)
            
            print(f"✅ Mock JSON email saved to: {filepath}")
            print(f"   Recipient: {recipient_email}")
            print(f"   Invoice: {image_filename}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error saving mock JSON email: {str(e)}")
            return False

    def send_invoice_excel(
        self,
        recipient_email: str,
        invoice_data_list: List[Dict[str, Any]],
        user_email: str
    ) -> bool:
        """Save all processed documents as Excel file instead of sending email"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"invoice_excel_{timestamp}_{recipient_email.replace('@', '_')}.xlsx"
            filepath = self.output_dir / filename
            
            # Create Excel file
            excel_data = self._create_excel_file(invoice_data_list)
            
            # Save Excel file
            with open(filepath, 'wb') as f:
                f.write(excel_data)
            
            print(f"✅ Mock Excel email saved to: {filepath}")
            print(f"   Recipient: {recipient_email}")
            print(f"   Documents: {len(invoice_data_list)} processed documents")
            print(f"   Sheets: Header (summary) + Details (line items)")
            
            return True
            
        except Exception as e:
            print(f"❌ Error saving mock Excel email: {str(e)}")
            return False

    def _create_excel_file(self, invoice_data_list: List[Dict[str, Any]]) -> bytes:
        """Create Excel file with Header and Details sheets (same as EmailService)"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Header sheet
        header_sheet = wb.create_sheet("Header")
        self._create_header_sheet(header_sheet, invoice_data_list)
        
        # Create Details sheet
        details_sheet = wb.create_sheet("Details")
        self._create_details_sheet(details_sheet, invoice_data_list)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()

    def _create_header_sheet(self, sheet, invoice_data_list: List[Dict[str, Any]]):
        """Create header sheet with summary information (same as EmailService)"""
        
        # Set up header row
        headers = [
            'Document ID', 'Filename', 'Processing Date', 'Invoice Number', 
            'Invoice Date', 'Due Date', 'Supplier Company', 'Customer Company',
            'Subtotal', 'Tax Amount', 'Total Amount', 'Processing Method', 
            'Confidence Score', 'Words Extracted', 'OCR Confidence'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row, invoice_data in enumerate(invoice_data_list, 2):
            structured_data = invoice_data.get('structured_data', {})
            document_info = structured_data.get('document_info', {})
            supplier = structured_data.get('supplier', {})
            customer = structured_data.get('customer', {})
            financial_details = structured_data.get('financial_details', {})
            ocr_stats = invoice_data.get('ocr_stats', {})
            
            # Fill data
            sheet.cell(row=row, column=1, value=invoice_data.get('image_id', ''))
            sheet.cell(row=row, column=2, value=invoice_data.get('original_filename', ''))
            sheet.cell(row=row, column=3, value=invoice_data.get('created_at', ''))
            sheet.cell(row=row, column=4, value=document_info.get('invoice_number', ''))
            sheet.cell(row=row, column=5, value=document_info.get('invoice_date', ''))
            sheet.cell(row=row, column=6, value=document_info.get('due_date', ''))
            sheet.cell(row=row, column=7, value=supplier.get('company_name', ''))
            sheet.cell(row=row, column=8, value=customer.get('company_name', ''))
            sheet.cell(row=row, column=9, value=financial_details.get('subtotal', 0))
            sheet.cell(row=row, column=10, value=financial_details.get('total_tax_amount', 0))
            sheet.cell(row=row, column=11, value=financial_details.get('total_after_tax', 0))
            sheet.cell(row=row, column=12, value=invoice_data.get('processing_method', ''))
            sheet.cell(row=row, column=13, value=invoice_data.get('extraction_confidence', 0))
            sheet.cell(row=row, column=14, value=ocr_stats.get('words_extracted', 0))
            sheet.cell(row=row, column=15, value=ocr_stats.get('average_confidence', 0))
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_details_sheet(self, sheet, invoice_data_list: List[Dict[str, Any]]):
        """Create details sheet with line items (same as EmailService)"""
        
        # Set up header row
        headers = [
            'Document ID', 'Filename', 'Invoice Number', 'Line Item #', 
            'Item Description', 'Quantity', 'Unit Price', 'Line Total',
            'Supplier Company', 'Customer Company', 'Invoice Date'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        current_row = 2
        for invoice_data in invoice_data_list:
            structured_data = invoice_data.get('structured_data', {})
            document_info = structured_data.get('document_info', {})
            supplier = structured_data.get('supplier', {})
            customer = structured_data.get('customer', {})
            line_items = structured_data.get('line_items', [])
            
            if line_items:
                for line_num, item in enumerate(line_items, 1):
                    sheet.cell(row=current_row, column=1, value=invoice_data.get('image_id', ''))
                    sheet.cell(row=current_row, column=2, value=invoice_data.get('original_filename', ''))
                    sheet.cell(row=current_row, column=3, value=document_info.get('invoice_number', ''))
                    sheet.cell(row=current_row, column=4, value=line_num)
                    sheet.cell(row=current_row, column=5, value=item.get('description', ''))
                    sheet.cell(row=current_row, column=6, value=item.get('quantity', 0))
                    sheet.cell(row=current_row, column=7, value=item.get('unit_price', 0))
                    sheet.cell(row=current_row, column=8, value=item.get('total_line_amount', 0))
                    sheet.cell(row=current_row, column=9, value=supplier.get('company_name', ''))
                    sheet.cell(row=current_row, column=10, value=customer.get('company_name', ''))
                    sheet.cell(row=current_row, column=11, value=document_info.get('invoice_date', ''))
                    current_row += 1
            else:
                # Add a row even if no line items
                sheet.cell(row=current_row, column=1, value=invoice_data.get('image_id', ''))
                sheet.cell(row=current_row, column=2, value=invoice_data.get('original_filename', ''))
                sheet.cell(row=current_row, column=3, value=document_info.get('invoice_number', ''))
                sheet.cell(row=current_row, column=4, value='N/A')
                sheet.cell(row=current_row, column=5, value='No line items found')
                sheet.cell(row=current_row, column=9, value=supplier.get('company_name', ''))
                sheet.cell(row=current_row, column=10, value=customer.get('company_name', ''))
                sheet.cell(row=current_row, column=11, value=document_info.get('invoice_date', ''))
                current_row += 1
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_invoice_html(self, invoice_data: Dict[str, Any], filename: str, user_email: str, recipient_email: str) -> str:
        """Create HTML formatted invoice email for mock service"""
        
        # Extract structured data
        structured_data = invoice_data.get('structured_data', {})
        document_info = structured_data.get('document_info', {})
        supplier = structured_data.get('supplier', {})
        customer = structured_data.get('customer', {})
        financial_details = structured_data.get('financial_details', {})
        line_items = structured_data.get('line_items', [])
        
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Invoice Data - {filename}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #333; }}
        .mock-notice {{ background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        .header {{ background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
        .section {{ margin-bottom: 25px; }}
        .section-title {{ background-color: #007bff; color: white; padding: 10px; margin: 0; font-size: 16px; }}
        .section-content {{ border: 1px solid #dee2e6; padding: 15px; }}
        .info-row {{ margin-bottom: 8px; }}
        .label {{ font-weight: bold; color: #495057; }}
        .value {{ margin-left: 10px; }}
        .line-items {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        .line-items th {{ background-color: #f8f9fa; padding: 8px; border: 1px solid #dee2e6; text-align: left; }}
        .line-items td {{ padding: 8px; border: 1px solid #dee2e6; }}
        .footer {{ background-color: #f8f9fa; padding: 15px; border-radius: 8px; margin-top: 20px; font-size: 12px; color: #666; }}
    </style>
</head>
<body>
    <div class="mock-notice">
        <strong>🧪 MOCK EMAIL SERVICE</strong><br>
        This email was generated by the mock email service for testing purposes.<br>
        In production, this would be sent to: <strong>{recipient_email}</strong>
    </div>

    <div class="header">
        <h1>Invoice Data Extraction</h1>
        <p><strong>File:</strong> {filename}</p>
        <p><strong>Processed by:</strong> {user_email}</p>
        <p><strong>Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
    </div>
"""

        # Document Information
        if document_info:
            html += """
    <div class="section">
        <h2 class="section-title">Document Information</h2>
        <div class="section-content">
"""
            if document_info.get('invoice_number'):
                html += f'<div class="info-row"><span class="label">Invoice Number:</span><span class="value">{document_info["invoice_number"]}</span></div>'
            if document_info.get('invoice_date'):
                html += f'<div class="info-row"><span class="label">Invoice Date:</span><span class="value">{document_info["invoice_date"]}</span></div>'
            if document_info.get('due_date'):
                html += f'<div class="info-row"><span class="label">Due Date:</span><span class="value">{document_info["due_date"]}</span></div>'
            html += "</div></div>"

        # Financial Summary (simplified for mock)
        if financial_details:
            html += """
    <div class="section">
        <h2 class="section-title">Financial Summary</h2>
        <div class="section-content">
"""
            if financial_details.get('subtotal'):
                html += f'<div class="info-row"><span class="label">Subtotal:</span><span class="value">${financial_details["subtotal"]:.2f}</span></div>'
            if financial_details.get('total_tax_amount'):
                html += f'<div class="info-row"><span class="label">Tax Amount:</span><span class="value">${financial_details["total_tax_amount"]:.2f}</span></div>'
            if financial_details.get('total_after_tax'):
                html += f'<div class="info-row"><span class="label"><strong>Total Amount:</strong></span><span class="value"><strong>${financial_details["total_after_tax"]:.2f}</strong></span></div>'
            html += "</div></div>"

        html += f"""
    <div class="footer">
        <p>This mock email was generated by CoreSRP invoice processing system.</p>
        <p>For support, please contact: noreply@coresrp.com</p>
        <p><small>Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</small></p>
    </div>
</body>
</html>
"""
        return html
    
    def send_email(self, recipient_email: str, subject: str, message: str) -> bool:
        """Send OTP email by saving to file (mock service)"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"otp_email_{timestamp}_{recipient_email.replace('@', '_')}.html"
            filepath = self.output_dir / filename
            
            html_content = self._create_otp_email_html(recipient_email, subject, message)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ Mock OTP email saved to: {filepath}")
            print(f"   Recipient: {recipient_email}")
            print(f"   Subject: {subject}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error saving mock OTP email: {str(e)}")
            return False
    
    def _create_otp_email_html(self, recipient_email: str, subject: str, message: str) -> str:
        """Create HTML formatted OTP email for mock service"""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{subject}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; color: #333; background-color: #f5f5f5; }}
        .container {{ max-width: 600px; margin: 0 auto; background-color: white; padding: 0; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .mock-notice {{ background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px 5px 0 0; margin: 0; }}
        .header {{ background-color: #4f46e5; color: white; padding: 30px; text-align: center; }}
        .content {{ padding: 30px; }}
        .footer {{ background-color: #f8f9fa; padding: 20px; border-radius: 0 0 8px 8px; font-size: 12px; color: #666; text-align: center; }}
        .otp-code {{ font-size: 32px; font-weight: bold; text-align: center; padding: 20px; background: #f8fafc; border-radius: 8px; margin: 20px 0; letter-spacing: 4px; }}
        .verification {{ color: #4f46e5; }}
        .password-reset {{ color: #dc2626; }}
        .login {{ color: #059669; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="mock-notice">
            <strong>🧪 MOCK EMAIL SERVICE</strong><br>
            This email was generated by the mock email service for testing purposes.<br>
            In production, this would be sent to: <strong>{recipient_email}</strong>
        </div>

        <div class="header">
            <h1>CoreSRP</h1>
            <p>{subject}</p>
        </div>
        
        <div class="content">
            {message}
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #eee;">
            <p><small><strong>This is a test email from the mock service.</strong><br>
            Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</small></p>
        </div>
        
        <div class="footer">
            <p>This is a mock email from CoreSRP for development/testing purposes.</p>
            <p>If this were production, this email would have been sent to {recipient_email}</p>
        </div>
    </div>
</body>
</html>
"""
        return html