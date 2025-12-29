"""
Import/Export API Endpoints

Provides template download, data export, and bulk import functionality
for Vendors (Address Book) and Item Master.
"""

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, Integer
from typing import Optional, List
from datetime import datetime
from decimal import Decimal
import io
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models import User, AddressBook, AddressBookContact, ItemMaster, ItemCategory, Warehouse
from app.api.auth import get_current_user

router = APIRouter()
logger = logging.getLogger(__name__)


# =============================================================================
# Helper Functions
# =============================================================================

def require_admin(user: User):
    """Ensure user has admin role"""
    if user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )


def create_styled_workbook(columns: List[str], sheet_name: str = "Data") -> Workbook:
    """Create a styled Excel workbook with headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Set column width based on header length
        ws.column_dimensions[cell.column_letter].width = max(15, len(column) + 5)

    # Freeze header row
    ws.freeze_panes = "A2"

    return wb


def get_next_address_number(db: Session, company_id: int) -> int:
    """Get the next address number as integer for company"""
    # Get max address number currently in use
    result = db.query(func.max(func.cast(AddressBook.address_number, Integer))).filter(
        AddressBook.company_id == company_id
    ).scalar()

    return (result or 0) + 1


def generate_address_number(number: int) -> str:
    """Format address number as 8-digit padded string"""
    return str(number).zfill(8)


# =============================================================================
# Template Downloads
# =============================================================================

VENDOR_COLUMNS = [
    "company_name",      # Required - Vendor/Supplier company name
    "tax_number",        # Tax registration number
    "address",           # Street address
    "city",              # City
    "country",           # Country
    "phone",             # Phone number
    "email",             # Email address
    "contact_person",    # Primary contact name
    "payment_terms",     # Payment terms (e.g., Net 30)
    "notes"              # Additional notes
]

ITEM_COLUMNS = [
    "item_number",       # Required - Unique item code
    "description",       # Required - Item description
    "category",          # Category name (will create if not exists)
    "unit",              # Unit of measure (e.g., EA, BOX, KG)
    "unit_cost",         # Purchase cost
    "unit_price",        # Selling price
    "minimum_stock",     # Minimum stock level
    "reorder_quantity",  # Reorder quantity
    "vendor_code",       # Vendor tax number for linking
    "notes"              # Additional notes
]


@router.get("/import-export/template/{entity_type}")
async def download_template(
    entity_type: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download an Excel template for the specified entity type.
    Supported types: vendors, items
    """
    require_admin(current_user)

    if entity_type == "vendors":
        columns = VENDOR_COLUMNS
        sheet_name = "Vendors"
        filename = "vendors_template.xlsx"
    elif entity_type == "items":
        columns = ITEM_COLUMNS
        sheet_name = "Items"
        filename = "items_template.xlsx"
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown entity type: {entity_type}. Supported types: vendors, items"
        )

    wb = create_styled_workbook(columns, sheet_name)

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Import Instructions"],
        [""],
        ["1. Fill in the data in the 'Data' sheet"],
        ["2. Required fields are marked in the column headers"],
        ["3. Do not modify the column headers"],
        ["4. Save the file and upload it in the Import section"],
        [""],
        ["Column Descriptions:"],
    ]

    if entity_type == "vendors":
        instructions.extend([
            ["company_name", "Required. The vendor/supplier company name"],
            ["tax_number", "Tax registration number (used for matching)"],
            ["address", "Street address"],
            ["city", "City name"],
            ["country", "Country name"],
            ["phone", "Phone number"],
            ["email", "Email address"],
            ["contact_person", "Primary contact person name"],
            ["payment_terms", "Payment terms (e.g., Net 30, Net 60)"],
            ["notes", "Additional notes or comments"],
        ])
    else:
        instructions.extend([
            ["item_number", "Required. Unique item code/SKU"],
            ["description", "Required. Item description"],
            ["category", "Category name (will be created if doesn't exist)"],
            ["unit", "Unit of measure (e.g., EA, BOX, KG, L)"],
            ["unit_cost", "Purchase cost per unit"],
            ["unit_price", "Selling price per unit"],
            ["minimum_stock", "Minimum stock level for reorder alerts"],
            ["reorder_quantity", "Default quantity to reorder"],
            ["vendor_code", "Vendor tax number for linking to supplier"],
            ["notes", "Additional notes or comments"],
        ])

    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_instructions.cell(row=row_idx, column=col_idx, value=value)

    # Style instruction header
    ws_instructions["A1"].font = Font(bold=True, size=14)
    ws_instructions.column_dimensions["A"].width = 20
    ws_instructions.column_dimensions["B"].width = 60

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# Data Export
# =============================================================================

@router.get("/import-export/export/{entity_type}")
async def export_data(
    entity_type: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export existing data to Excel format.
    Supported types: vendors, items
    """
    require_admin(current_user)

    if entity_type == "vendors":
        return await export_vendors(db, current_user)
    elif entity_type == "items":
        return await export_items(db, current_user)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown entity type: {entity_type}. Supported types: vendors, items"
        )


async def export_vendors(db: Session, current_user: User) -> StreamingResponse:
    """Export vendors to Excel"""
    vendors = db.query(AddressBook).filter(
        AddressBook.company_id == current_user.company_id,
        AddressBook.search_type == "V"  # V = Vendor
    ).order_by(AddressBook.alpha_name).all()

    wb = create_styled_workbook(VENDOR_COLUMNS, "Vendors")
    ws = wb.active

    for row_idx, vendor in enumerate(vendors, 2):
        # Get primary contact if exists
        primary_contact = None
        if vendor.contacts:
            primary_contact = next((c for c in vendor.contacts if c.is_primary), None) or (vendor.contacts[0] if vendor.contacts else None)

        ws.cell(row=row_idx, column=1, value=vendor.alpha_name)
        ws.cell(row=row_idx, column=2, value=vendor.tax_id)
        ws.cell(row=row_idx, column=3, value=vendor.address_line_1)
        ws.cell(row=row_idx, column=4, value=vendor.city)
        ws.cell(row=row_idx, column=5, value=vendor.country)
        ws.cell(row=row_idx, column=6, value=vendor.phone_primary)
        ws.cell(row=row_idx, column=7, value=vendor.email)
        ws.cell(row=row_idx, column=8, value=primary_contact.full_name if primary_contact else None)
        ws.cell(row=row_idx, column=9, value=vendor.category_code_04)  # Payment terms stored in category code
        ws.cell(row=row_idx, column=10, value=vendor.notes)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"vendors_export_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def export_items(db: Session, current_user: User) -> StreamingResponse:
    """Export items to Excel"""
    items = db.query(ItemMaster).filter(
        ItemMaster.company_id == current_user.company_id
    ).order_by(ItemMaster.item_number).all()

    # Get categories for mapping
    categories = {c.id: c.name for c in db.query(ItemCategory).filter(
        ItemCategory.company_id == current_user.company_id
    ).all()}

    # Get vendors for mapping
    vendors = {v.id: v.tax_id for v in db.query(AddressBook).filter(
        AddressBook.company_id == current_user.company_id,
        AddressBook.search_type == "V"
    ).all()}

    wb = create_styled_workbook(ITEM_COLUMNS, "Items")
    ws = wb.active

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.item_number)
        ws.cell(row=row_idx, column=2, value=item.description)
        ws.cell(row=row_idx, column=3, value=categories.get(item.category_id, ""))
        ws.cell(row=row_idx, column=4, value=item.unit)
        ws.cell(row=row_idx, column=5, value=float(item.unit_cost) if item.unit_cost else None)
        ws.cell(row=row_idx, column=6, value=float(item.unit_price) if item.unit_price else None)
        ws.cell(row=row_idx, column=7, value=item.minimum_stock_level)
        ws.cell(row=row_idx, column=8, value=item.reorder_quantity)
        ws.cell(row=row_idx, column=9, value=vendors.get(item.primary_address_book_id, ""))
        ws.cell(row=row_idx, column=10, value=item.notes)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"items_export_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# Data Import
# =============================================================================

@router.post("/import-export/import/{entity_type}")
async def import_data(
    entity_type: str,
    file: UploadFile = File(...),
    skip_duplicates: bool = Form(True),
    update_existing: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Import data from Excel file.
    Supported types: vendors, items
    """
    require_admin(current_user)

    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload an Excel (.xlsx, .xls) or CSV file."
        )

    try:
        # Read file content
        content = await file.read()

        if file.filename.endswith('.csv'):
            # Handle CSV
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV import not yet supported. Please use Excel format (.xlsx)"
            )

        # Load Excel workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        if entity_type == "vendors":
            return await import_vendors(db, current_user, ws, skip_duplicates, update_existing)
        elif entity_type == "items":
            return await import_items(db, current_user, ws, skip_duplicates, update_existing)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown entity type: {entity_type}. Supported types: vendors, items"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process file: {str(e)}"
        )


async def import_vendors(
    db: Session,
    current_user: User,
    ws,
    skip_duplicates: bool,
    update_existing: bool
) -> dict:
    """Import vendors from worksheet"""
    created = 0
    updated = 0
    skipped = 0
    errors = []

    # Get header row
    headers = [cell.value for cell in ws[1] if cell.value]

    # Validate headers
    expected_headers = set(VENDOR_COLUMNS[:2])  # At least company_name
    found_headers = set(h.lower().replace(" ", "_") if h else "" for h in headers)

    if "company_name" not in found_headers:
        return {
            "success": False,
            "message": "Missing required column: company_name",
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": ["The 'company_name' column is required"]
        }

    # Map column indices
    col_map = {}
    for idx, header in enumerate(headers):
        if header:
            col_map[header.lower().replace(" ", "_")] = idx

    # Get starting address number for new records
    next_address_num = get_next_address_number(db, current_user.company_id)

    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            company_name = row[col_map.get("company_name", 0)] if col_map.get("company_name") is not None else None

            if not company_name:
                continue  # Skip empty rows

            tax_number = row[col_map.get("tax_number")] if col_map.get("tax_number") is not None else None

            # Check for existing vendor by tax_number or name
            existing = None
            if tax_number:
                existing = db.query(AddressBook).filter(
                    AddressBook.company_id == current_user.company_id,
                    AddressBook.search_type == "V",
                    AddressBook.tax_id == str(tax_number)
                ).first()

            if not existing:
                existing = db.query(AddressBook).filter(
                    AddressBook.company_id == current_user.company_id,
                    AddressBook.search_type == "V",
                    func.lower(AddressBook.alpha_name) == func.lower(str(company_name))
                ).first()

            if existing:
                if update_existing:
                    # Update existing vendor
                    if col_map.get("address") is not None and row[col_map["address"]]:
                        existing.address_line_1 = str(row[col_map["address"]])
                    if col_map.get("city") is not None and row[col_map["city"]]:
                        existing.city = str(row[col_map["city"]])
                    if col_map.get("country") is not None and row[col_map["country"]]:
                        existing.country = str(row[col_map["country"]])
                    if col_map.get("phone") is not None and row[col_map["phone"]]:
                        existing.phone_primary = str(row[col_map["phone"]])
                    if col_map.get("email") is not None and row[col_map["email"]]:
                        existing.email = str(row[col_map["email"]])
                    if col_map.get("payment_terms") is not None and row[col_map["payment_terms"]]:
                        existing.category_code_04 = str(row[col_map["payment_terms"]])
                    if col_map.get("notes") is not None and row[col_map["notes"]]:
                        existing.notes = str(row[col_map["notes"]])
                    if tax_number:
                        existing.tax_id = str(tax_number)
                    existing.updated_at = datetime.utcnow()
                    updated += 1
                elif skip_duplicates:
                    skipped += 1
                else:
                    errors.append(f"Row {row_idx}: Vendor '{company_name}' already exists")
                continue

            # Create new vendor
            address_number = generate_address_number(next_address_num)
            next_address_num += 1  # Increment for next record

            vendor = AddressBook(
                company_id=current_user.company_id,
                address_number=address_number,
                search_type="V",
                alpha_name=str(company_name),
                tax_id=str(tax_number) if tax_number else None,
                address_line_1=str(row[col_map["address"]]) if col_map.get("address") is not None and row[col_map["address"]] else None,
                city=str(row[col_map["city"]]) if col_map.get("city") is not None and row[col_map["city"]] else None,
                country=str(row[col_map["country"]]) if col_map.get("country") is not None and row[col_map["country"]] else None,
                phone_primary=str(row[col_map["phone"]]) if col_map.get("phone") is not None and row[col_map["phone"]] else None,
                email=str(row[col_map["email"]]) if col_map.get("email") is not None and row[col_map["email"]] else None,
                category_code_04=str(row[col_map["payment_terms"]]) if col_map.get("payment_terms") is not None and row[col_map["payment_terms"]] else None,
                notes=str(row[col_map["notes"]]) if col_map.get("notes") is not None and row[col_map["notes"]] else None,
                is_active=True,
                created_by=current_user.id
            )

            db.add(vendor)

            # Create contact person if provided
            contact_person_name = row[col_map["contact_person"]] if col_map.get("contact_person") is not None and row[col_map["contact_person"]] else None
            if contact_person_name:
                db.flush()  # Get vendor ID
                contact = AddressBookContact(
                    address_book_id=vendor.id,
                    line_number=1,
                    full_name=str(contact_person_name),
                    contact_type="primary",
                    is_primary=True,
                    is_active=True
                )
                db.add(contact)

            created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    db.commit()

    total = created + updated + skipped
    success = len(errors) == 0 or (created + updated) > 0

    return {
        "success": success,
        "message": f"Import completed. {created} created, {updated} updated, {skipped} skipped." + (f" {len(errors)} errors." if errors else ""),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10]  # Limit errors to first 10
    }


async def import_items(
    db: Session,
    current_user: User,
    ws,
    skip_duplicates: bool,
    update_existing: bool
) -> dict:
    """Import items from worksheet"""
    created = 0
    updated = 0
    skipped = 0
    errors = []

    # Get header row
    headers = [cell.value for cell in ws[1] if cell.value]

    # Validate headers
    found_headers = set(h.lower().replace(" ", "_") if h else "" for h in headers)

    if "item_number" not in found_headers or "description" not in found_headers:
        return {
            "success": False,
            "message": "Missing required columns: item_number and description are required",
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": ["The 'item_number' and 'description' columns are required"]
        }

    # Map column indices
    col_map = {}
    for idx, header in enumerate(headers):
        if header:
            col_map[header.lower().replace(" ", "_")] = idx

    # Cache categories and vendors - cache by both name and code
    all_categories = db.query(ItemCategory).filter(
        ItemCategory.company_id == current_user.company_id
    ).all()
    categories_by_name = {c.name.lower(): c for c in all_categories}
    categories_by_code = {c.code.upper(): c for c in all_categories}

    vendors = {v.tax_id.lower(): v for v in db.query(AddressBook).filter(
        AddressBook.company_id == current_user.company_id,
        AddressBook.search_type == "V",
        AddressBook.tax_id.isnot(None)
    ).all() if v.tax_id}

    # Process rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            item_number = row[col_map.get("item_number", 0)] if col_map.get("item_number") is not None else None
            description = row[col_map.get("description", 1)] if col_map.get("description") is not None else None

            if not item_number or not description:
                continue  # Skip empty rows

            item_number = str(item_number).strip()

            # Check for existing item
            existing = db.query(ItemMaster).filter(
                ItemMaster.company_id == current_user.company_id,
                func.lower(ItemMaster.item_number) == func.lower(item_number)
            ).first()

            if existing:
                if update_existing:
                    # Update existing item
                    existing.description = str(description)

                    # Update category if provided
                    if col_map.get("category") is not None and row[col_map["category"]]:
                        cat_name = str(row[col_map["category"]]).strip().lower()
                        cat_name_str = str(row[col_map["category"]]).strip()
                        cat_code = cat_name_str[:10].upper().replace(" ", "")

                        if cat_name in categories_by_name:
                            existing.category_id = categories_by_name[cat_name].id
                        elif cat_code in categories_by_code:
                            # Category with this code exists, use it
                            existing.category_id = categories_by_code[cat_code].id
                            categories_by_name[cat_name] = categories_by_code[cat_code]
                        else:
                            # Create new category
                            new_cat = ItemCategory(
                                company_id=current_user.company_id,
                                code=cat_code,
                                name=cat_name_str
                            )
                            db.add(new_cat)
                            db.flush()
                            categories_by_name[cat_name] = new_cat
                            categories_by_code[cat_code] = new_cat
                            existing.category_id = new_cat.id

                    if col_map.get("unit") is not None and row[col_map["unit"]]:
                        existing.unit = str(row[col_map["unit"]])
                    if col_map.get("unit_cost") is not None and row[col_map["unit_cost"]]:
                        existing.unit_cost = Decimal(str(row[col_map["unit_cost"]]))
                    if col_map.get("unit_price") is not None and row[col_map["unit_price"]]:
                        existing.unit_price = Decimal(str(row[col_map["unit_price"]]))
                    if col_map.get("minimum_stock") is not None and row[col_map["minimum_stock"]]:
                        existing.minimum_stock_level = int(row[col_map["minimum_stock"]])
                    if col_map.get("reorder_quantity") is not None and row[col_map["reorder_quantity"]]:
                        existing.reorder_quantity = int(row[col_map["reorder_quantity"]])
                    if col_map.get("notes") is not None and row[col_map["notes"]]:
                        existing.notes = str(row[col_map["notes"]])

                    # Update vendor if provided
                    if col_map.get("vendor_code") is not None and row[col_map["vendor_code"]]:
                        vendor_code = str(row[col_map["vendor_code"]]).strip().lower()
                        if vendor_code in vendors:
                            existing.primary_address_book_id = vendors[vendor_code].id

                    existing.updated_at = datetime.utcnow()
                    updated += 1
                elif skip_duplicates:
                    skipped += 1
                else:
                    errors.append(f"Row {row_idx}: Item '{item_number}' already exists")
                continue

            # Process category
            category_id = None
            if col_map.get("category") is not None and row[col_map["category"]]:
                cat_name = str(row[col_map["category"]]).strip().lower()
                cat_name_str = str(row[col_map["category"]]).strip()
                cat_code = cat_name_str[:10].upper().replace(" ", "")

                if cat_name in categories_by_name:
                    category_id = categories_by_name[cat_name].id
                elif cat_code in categories_by_code:
                    # Category with this code exists, use it
                    category_id = categories_by_code[cat_code].id
                    categories_by_name[cat_name] = categories_by_code[cat_code]
                else:
                    # Create new category
                    new_cat = ItemCategory(
                        company_id=current_user.company_id,
                        code=cat_code,
                        name=cat_name_str
                    )
                    db.add(new_cat)
                    db.flush()
                    categories_by_name[cat_name] = new_cat
                    categories_by_code[cat_code] = new_cat
                    category_id = new_cat.id

            # Process vendor
            vendor_id = None
            if col_map.get("vendor_code") is not None and row[col_map["vendor_code"]]:
                vendor_code = str(row[col_map["vendor_code"]]).strip().lower()
                if vendor_code in vendors:
                    vendor_id = vendors[vendor_code].id

            # Create new item
            item = ItemMaster(
                company_id=current_user.company_id,
                item_number=item_number,
                description=str(description),
                category_id=category_id,
                unit=str(row[col_map["unit"]]) if col_map.get("unit") is not None and row[col_map["unit"]] else "EA",
                unit_cost=Decimal(str(row[col_map["unit_cost"]])) if col_map.get("unit_cost") is not None and row[col_map["unit_cost"]] else Decimal("0"),
                unit_price=Decimal(str(row[col_map["unit_price"]])) if col_map.get("unit_price") is not None and row[col_map["unit_price"]] else Decimal("0"),
                minimum_stock_level=int(row[col_map["minimum_stock"]]) if col_map.get("minimum_stock") is not None and row[col_map["minimum_stock"]] else 0,
                reorder_quantity=int(row[col_map["reorder_quantity"]]) if col_map.get("reorder_quantity") is not None and row[col_map["reorder_quantity"]] else 0,
                primary_address_book_id=vendor_id,
                notes=str(row[col_map["notes"]]) if col_map.get("notes") is not None and row[col_map["notes"]] else None,
                is_active=True,
                created_by=current_user.id
            )

            db.add(item)
            created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    db.commit()

    total = created + updated + skipped
    success = len(errors) == 0 or (created + updated) > 0

    return {
        "success": success,
        "message": f"Import completed. {created} created, {updated} updated, {skipped} skipped." + (f" {len(errors)} errors." if errors else ""),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10]  # Limit errors to first 10
    }
